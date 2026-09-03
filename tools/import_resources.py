#!/usr/bin/env python3
"""Convert the YuBan resource workbook into the website JSON index."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


REQUIRED_HEADERS = (
    "資源名稱",
    "連結",
    "摘要",
    "內容類型",
    "提供方／來源類型",
    "原始來源",
    "是否為入口型資源",
    "年齡階段",
    "主題",
    "關鍵標籤",
    "審核狀態",
    "可信度備註",
    "注意事項",
    "年齡群組",
    "地區",
    "資源類型",
    "使用對象",
    "來源地區",
    "語言",
)
LEGACY_REQUIRED_HEADERS = REQUIRED_HEADERS[:13]

TYPE_MAP = {
    "文章": "文章",
    "影片": "影片",
    "工具／用品": "工具／用品",
    "混合型內容": "混合型內容",
    "連結入口": "連結入口",
    "網站入口": "連結入口",
    "Podcast入口": "連結入口",
    "分類入口": "連結入口",
    "課程平台": "連結入口",
    "工具平台": "工具／用品",
    "搜尋工具": "工具／用品",
    "互動工具": "工具／用品",
    "文章／工具": "混合型內容",
    "圖文文章": "文章",
    "YouTube影片": "影片",
    "電子手冊": "文章",
    "資訊專區／文件下載": "文章",
    "政策／服務說明": "文章",
    "政策專頁": "文章",
    "衛教文章／門診資訊": "文章",
    "圖表／衛教說明": "文章",
    "影音／宣導教材": "影片",
    "影音／衛教專區": "混合型內容",
    "查詢工具": "工具／用品",
    "保護專線／線上服務": "工具／用品",
    "網安資源／申訴入口": "工具／用品",
    "連結入口": "連結入口",
    "資源入口": "連結入口",
    "線上課程入口": "連結入口",
    "查詢入口": "連結入口",
    "醫院衛教專區": "連結入口",
    "專業衛教專區": "連結入口",
    "資訊入口": "連結入口",
    "專業資訊平台": "連結入口",
    "專業資源入口": "連結入口",
    "親職內容專區": "連結入口",
}

AGE_STAGE_MAP = {
    "全齡": ["全齡"],
    "孕前–2歲": ["孕期", "0-1歲", "1-3歲"],
    "孕前–6歲": ["孕期", "0-1歲", "1-3歲", "3-6歲"],
    "孕期–2歲以上": ["孕期", "0-1歲", "1-3歲", "3-6歲"],
    "孕期–學齡前": ["孕期", "0-1歲", "1-3歲", "3-6歲"],
    "孕期–6歲": ["孕期", "0-1歲", "1-3歲", "3-6歲"],
    "0–1歲": ["0-1歲"],
    "0–3歲": ["0-1歲", "1-3歲"],
    "0–6歲": ["0-1歲", "1-3歲", "3-6歲"],
    "0–7歲": ["0-1歲", "1-3歲", "3-6歲"],
    "0–18歲": ["0-1歲", "1-3歲", "3-6歲"],
    "0–18歲及照顧者": ["0-1歲", "1-3歲", "3-6歲"],
    "0歲–學齡前": ["0-1歲", "1-3歲", "3-6歲"],
    "2–6歲": ["1-3歲", "3-6歲"],
    "6–12歲": ["3-6歲"],
    "6–18歲及照顧者": ["3-6歲"],
    "0–15歲": ["0-1歲", "1-3歲", "3-6歲", "全齡"],
    "全年齡": ["全齡"],
    "嬰兒": ["0-1歲"],
    "嬰兒至學齡前": ["0-1歲", "1-3歲", "3-6歲"],
    "嬰幼兒": ["0-1歲", "1-3歲"],
    "嬰幼兒／學齡前": ["0-1歲", "1-3歲", "3-6歲"],
    "孕期至學齡前": ["孕期", "0-1歲", "1-3歲", "3-6歲"],
    "孕 期至學齡前": ["孕期", "0-1歲", "1-3歲", "3-6歲"],
    "孕期至幼兒期": ["孕期", "0-1歲", "1-3歲"],
    "孕期／新生兒": ["孕期", "0-1歲"],
    "學步兒": ["1-3歲"],
    "學齡前": ["3-6歲"],
    "家長與兒童": ["全齡"],
    "幼兒至學齡兒童": ["1-3歲", "3-6歲"],
    "幼兒／學齡兒童": ["1-3歲", "3-6歲"],
    "新生兒／嬰兒": ["0-1歲"],
}

TOPIC_GROUPS = {
    "綜合育兒": "親職與家庭",
    "健康與安全": "健康與照護",
    "健康與發展": "健康與照護",
    "學習與教育": "發展與學習",
    "學習與閱讀": "發展與學習",
    "情緒與行為": "情緒與心理",
    "托育與幼兒園": "托育與服務",
    "托育 與幼兒園": "托育與服務",
    "發展與早療": "發展與學習",
    "睡眠": "健康與照護",
    "補助與權益": "親職與家庭",
    "親子互動與教養": "親職與家庭",
    "親子活動": "親職與家庭",
    "遊戲與學習": "發展與學習",
    "飲食與營養": "健康與照護",
    "孕產與嬰幼兒照護": "健康與照護",
    "健康與生活照護": "健康與照護",
    "預防接種": "健康與照護",
    "兒童預防保健": "健康與照護",
    "新生兒照護": "健康與照護",
    "兒童健康": "健康與照護",
    "母乳哺育": "健康與照護",
    "兒童就醫資源": "健康與照護",
    "生長發育": "健康與照護",
    "視力保健": "健康與照護",
    "兒童口腔保健": "健康與照護",
    "兒科疾病與照護": "健康與照護",
    "兒科疾病衛教": "健康與照護",
    "早產兒照護": "健康與照護",
    "兒童發展篩檢": "發展與學習",
    "早期療育": "發展與學習",
    "幼兒園與學前教保": "發展與學習",
    "課後照顧": "發展與學習",
    "兒童發展與早療": "發展與學習",
    "親子共讀": "發展與學習",
    "親職知能": "親職與家庭",
    "補助與家庭支持": "親職與家庭",
    "親職與家庭關係": "親職與家庭",
    "親職教養與家庭支持": "親職與家庭",
    "兒童事故傷害預防": "安全與保護",
    "兒少保護": "安全與保護",
    "兒童權利": "安全與保護",
    "兒少網路安全": "安全與保護",
    "托育媒合": "托育與服務",
    "兒童青少年心理健康": "情緒與心理",
}

AGE_GROUP_MAP = {
    "0–1歲": "學齡前",
    "0–3歲": "學齡前",
    "0–5歲": "學齡前",
    "0–6歲": "學齡前",
    "0–12歲": "國小",
    "0–15歲": "全齡",
    "2–3歲": "學齡前",
    "2–6歲": "學齡前",
    "2–12歲": "國小",
    "3–5歲": "學齡前",
    "孕期–0歲": "學齡前",
    "孕期–3歲": "學齡前",
    "孕期–6歲": "學齡前",
}

RESOURCE_CATEGORY_MAP = {
    "入口型資源": "機構",
    "工具型": "機構",
    "互動工具型": "機構",
    "分齡指南型": "學習教材",
    "問答型": "學習教材",
    "地區活動指南型": "活動",
    "地方補助整理型": "補助",
    "實作指南型": "學習教材",
    "專業指南型": "學習教材",
    "影音實作型": "學習教材",
    "影音解說型": "學習教材",
    "影音音訊型": "學習教材",
    "情境解決型": "學習教材",
    "懶人包型": "學習教材",
    "指南型": "學習教材",
    "時效活動型": "活動",
    "權益指南型": "政策",
    "比較指南型": "學習教材",
    "決策指南型": "學習教材",
    "深度解讀型": "學習教材",
    "總覽型": "機構",
    "補助指南型": "補助",
    "課程型": "課程",
    "費用指南型": "學習教材",
    "選擇指南型": "學習教材",
}

AUDIENCE_MAP = {
    "家庭": ["家長"],
    "家長與幼兒": ["家長", "兒童"],
    "家長／兒童": ["家長", "兒童"],
    "家長／準父母": ["家長"],
    "家長／照顧者": ["家長"],
}


def split_tags(value: str) -> list[str]:
    return [normalize_text(tag) for tag in re.split(r"[,，、;；]", value) if tag.strip()]


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value)).strip()


def make_resource_id(url: str) -> str:
    digest = hashlib.sha256(url.encode("utf-8")).hexdigest()[:12]
    return f"resource-{digest}"


def parse_review_status(value: str) -> tuple[str, str | None]:
    value = normalize_text(value)
    reviewed_at = None
    date_match = re.search(r"(\d{4}-\d{2}-\d{2})", value)
    if date_match:
        reviewed_at = date_match.group(1)
    status = "verified" if value.startswith(("人工核實", "通過")) else "ai_draft"
    return status, reviewed_at


def validate_url(url: str, row_number: int) -> None:
    parsed = urlparse(url)
    if parsed.scheme not in {"http", "https"} or not parsed.netloc:
        raise ValueError(f"第 {row_number} 列不是有效的 HTTP(S) 網址：{url}")


def convert_row(row: dict[str, str], row_number: int) -> dict[str, object]:
    raw_type = normalize_text(row["內容類型"])
    age_label = normalize_text(row["年齡階段"])
    topic = normalize_text(row["主題"])
    url = normalize_text(row["連結"])

    if raw_type not in TYPE_MAP:
        raise ValueError(f"第 {row_number} 列有未定義的內容類型：{raw_type}")
    if age_label not in AGE_STAGE_MAP:
        raise ValueError(f"第 {row_number} 列有未定義的年齡階段：{age_label}")
    if topic not in TOPIC_GROUPS:
        raise ValueError(f"第 {row_number} 列有未定義的主題：{topic}")

    validate_url(url, row_number)
    review_status, reviewed_at = parse_review_status(row["審核狀態"])

    return {
        "id": make_resource_id(url),
        "title": row["資源名稱"],
        "summary": row["摘要"],
        "type": TYPE_MAP[raw_type],
        "type_label": raw_type,
        "source": {
            "name": row["原始來源"],
            "type": row["提供方／來源類型"],
            "url": url,
        },
        "is_hub": row["是否為入口型資源"] == "是",
        "age_label": age_label,
        "age_ranges": AGE_STAGE_MAP[age_label],
        "age_groups": [AGE_GROUP_MAP.get(value, value) for value in split_tags(row["年齡群組"])],
        "regions": split_tags(row["地區"]),
        "resource_categories": [RESOURCE_CATEGORY_MAP.get(value, value) for value in split_tags(row["資源類型"])],
        "audiences": [audience for value in split_tags(row["使用對象"]) for audience in AUDIENCE_MAP.get(value, [value])],
        "origin_region": normalize_text(row["來源地區"]),
        "languages": split_tags(row["語言"]),
        "topic": topic,
        "topic_group": TOPIC_GROUPS[topic],
        "tags": split_tags(row["關鍵標籤"]),
        "ai_summary_status": review_status,
        "reviewed_at": reviewed_at,
        "credibility_note": row["可信度備註"],
        "notice": row["注意事項"],
        "url": url,
    }


def load_resources(workbook_path: Path) -> list[dict[str, object]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)

    try:
        headers = tuple(re.sub(r"\s+", "", str(value)) if value is not None else "" for value in next(rows))
    except StopIteration as exc:
        raise ValueError("工作簿沒有任何資料") from exc

    if headers not in (REQUIRED_HEADERS, LEGACY_REQUIRED_HEADERS):
        raise ValueError(
            "工作簿欄位與預期不符。\n"
            f"預期：{REQUIRED_HEADERS}\n"
            f"實際：{headers}"
        )

    resources: list[dict[str, object]] = []
    for row_number, values in enumerate(rows, start=2):
        if all(value in (None, "") for value in values):
            continue
        if any(value in (None, "") for value in values):
            missing = [headers[index] for index, value in enumerate(values) if value in (None, "")]
            raise ValueError(f"第 {row_number} 列缺少欄位：{', '.join(missing)}")

        row = {headers[index]: normalize_text(value) for index, value in enumerate(values)}
        if headers == LEGACY_REQUIRED_HEADERS:
            row.update(
                {
                    "年齡群組": "全齡" if row["年齡階段"] == "全齡" else "學齡前",
                    "地區": "全國",
                    "資源類型": "學習教材",
                    "使用對象": "家長",
                    "來源地區": "台灣",
                    "語言": "繁體中文",
                }
            )
        resources.append(convert_row(row, row_number))

    ids = [resource["id"] for resource in resources]
    urls = [resource["url"] for resource in resources]
    if len(ids) != len(set(ids)) or len(urls) != len(set(urls)):
        raise ValueError("工作簿包含重複網址")
    return resources


def merge_resources(
    existing: list[dict[str, object]], incoming: list[dict[str, object]]
) -> tuple[list[dict[str, object]], int]:
    existing_urls = {resource["url"] for resource in existing}
    existing_ids = {resource["id"] for resource in existing}
    merged = [*existing]
    skipped = 0

    for resource in incoming:
        if resource["url"] in existing_urls or resource["id"] in existing_ids:
            skipped += 1
            continue
        merged.append(resource)
        existing_urls.add(resource["url"])
        existing_ids.add(resource["id"])

    return merged, skipped


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("docs/resource_v0.2.xlsx"),
        help="來源 Excel 工作簿",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("website/data/sample-resources.json"),
        help="網站 JSON 輸出位置",
    )
    args = parser.parse_args()

    incoming = load_resources(args.input)
    existing = []
    if args.output.exists():
        existing = json.loads(args.output.read_text(encoding="utf-8"))
        if not isinstance(existing, list):
            raise ValueError("既有資源索引必須是 JSON 陣列")

    resources, skipped = merge_resources(existing, incoming)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(resources, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(
        f"Imported {len(incoming) - skipped} new resources into {args.output} "
        f"({skipped} duplicates skipped, {len(resources)} total)"
    )


if __name__ == "__main__":
    main()
